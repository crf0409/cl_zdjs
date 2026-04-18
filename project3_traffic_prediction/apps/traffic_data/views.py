"""交通数据管理模块 - 视图函数"""
import csv
import io
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Avg, Sum, Count, Q, Max, Min
from django.utils import timezone

from .models import TrafficFlow, WeatherData, TimePeriodLabel, PredictionResult
from apps.users.views import _log_activity


@login_required
def data_list_view(request):
    """交通数据列表视图
    展示所有交通流量数据，支持分页、搜索和多条件过滤
    GET参数: location, source, date_from, date_to, is_cleaned, page
    """
    queryset = TrafficFlow.objects.all()

    # 按监测点位过滤
    location = request.GET.get('location', '')
    if location:
        queryset = queryset.filter(location__icontains=location)

    # 按数据源过滤
    source = request.GET.get('source', '')
    if source:
        queryset = queryset.filter(source=source)

    # 按日期范围过滤
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        try:
            queryset = queryset.filter(timestamp__gte=datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            queryset = queryset.filter(timestamp__lte=datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    # 按清洗状态过滤
    is_cleaned = request.GET.get('is_cleaned', '')
    if is_cleaned in ('true', 'false'):
        queryset = queryset.filter(is_cleaned=(is_cleaned == 'true'))

    # 统计信息
    stats = queryset.aggregate(
        total_count=Count('id'),
        avg_flow=Avg('total_flow'),
        max_flow=Max('total_flow'),
        min_flow=Min('total_flow'),
        avg_speed=Avg('avg_speed'),
    )

    # 分页
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 获取所有监测点位和数据源选项（用于过滤下拉框）
    locations = TrafficFlow.objects.values_list('location', flat=True).distinct()[:50]

    context = {
        'page_obj': page_obj,
        'stats': stats,
        'locations': locations,
        'source_choices': TrafficFlow.SOURCE_CHOICES,
        # 保留过滤条件
        'location': location,
        'source': source,
        'date_from': date_from,
        'date_to': date_to,
        'is_cleaned': is_cleaned,
    }
    return render(request, 'traffic_data/data_list.html', context)


@login_required
def data_upload_view(request):
    """数据上传视图
    GET: 显示上传表单
    POST: 处理CSV/Excel文件上传，解析并存储交通数据
    支持的CSV列: timestamp, location, camera_id, vehicle_count, pedestrian_count,
                 bicycle_count, motorcycle_count, truck_count, avg_speed, occupancy_rate, source
    """
    if request.method == 'POST':
        uploaded_file = request.FILES.get('data_file')
        if not uploaded_file:
            messages.error(request, '请选择要上传的文件。')
            return redirect('data_upload')

        file_name = uploaded_file.name.lower()

        # 检查文件格式
        if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            messages.error(request, '仅支持CSV和Excel格式文件。')
            return redirect('data_upload')

        success_count = 0
        error_count = 0
        errors = []

        try:
            if file_name.endswith('.csv'):
                # 处理CSV文件
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded_file))

                for row_num, row in enumerate(reader, start=2):
                    try:
                        traffic_data = TrafficFlow(
                            timestamp=row.get('timestamp', '').strip(),
                            location=row.get('location', '').strip(),
                            camera_id=row.get('camera_id', '').strip(),
                            vehicle_count=int(row.get('vehicle_count', 0) or 0),
                            pedestrian_count=int(row.get('pedestrian_count', 0) or 0),
                            bicycle_count=int(row.get('bicycle_count', 0) or 0),
                            motorcycle_count=int(row.get('motorcycle_count', 0) or 0),
                            truck_count=int(row.get('truck_count', 0) or 0),
                            avg_speed=float(row.get('avg_speed', 0) or 0) if row.get('avg_speed') else None,
                            occupancy_rate=float(row.get('occupancy_rate', 0) or 0) if row.get('occupancy_rate') else None,
                            source=row.get('source', 'manual').strip() or 'manual',
                            uploaded_by=request.user,
                        )
                        traffic_data.save()
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f'第{row_num}行: {str(e)}')
                        if error_count >= 10:
                            errors.append('错误过多，已停止显示后续错误...')
                            break

            else:
                # 处理Excel文件
                try:
                    import openpyxl
                except ImportError:
                    messages.error(request, '服务器缺少openpyxl库，无法处理Excel文件。请使用CSV格式上传。')
                    return redirect('data_upload')

                wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        traffic_data = TrafficFlow(
                            timestamp=row_dict.get('timestamp', ''),
                            location=str(row_dict.get('location', '')).strip(),
                            camera_id=str(row_dict.get('camera_id', '')).strip(),
                            vehicle_count=int(row_dict.get('vehicle_count', 0) or 0),
                            pedestrian_count=int(row_dict.get('pedestrian_count', 0) or 0),
                            bicycle_count=int(row_dict.get('bicycle_count', 0) or 0),
                            motorcycle_count=int(row_dict.get('motorcycle_count', 0) or 0),
                            truck_count=int(row_dict.get('truck_count', 0) or 0),
                            avg_speed=float(row_dict.get('avg_speed', 0) or 0) if row_dict.get('avg_speed') else None,
                            occupancy_rate=float(row_dict.get('occupancy_rate', 0) or 0) if row_dict.get('occupancy_rate') else None,
                            source=str(row_dict.get('source', 'manual')).strip() or 'manual',
                            uploaded_by=request.user,
                        )
                        traffic_data.save()
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f'第{row_num}行: {str(e)}')
                        if error_count >= 10:
                            errors.append('错误过多，已停止显示后续错误...')
                            break

        except Exception as e:
            messages.error(request, f'文件解析失败: {str(e)}')
            return redirect('data_upload')

        # 记录用户上传活动
        _log_activity(request.user, 'upload', f'上传数据文件: {uploaded_file.name}，成功{success_count}条，失败{error_count}条', request)

        if success_count > 0:
            messages.success(request, f'数据上传成功！共导入 {success_count} 条记录。')
        if error_count > 0:
            messages.warning(request, f'有 {error_count} 条记录导入失败。')

        return render(request, 'traffic_data/data_upload.html', {
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10],
        })

    return render(request, 'traffic_data/data_upload.html')


@login_required
def data_detail_view(request, pk):
    """数据详情视图
    查看单条交通数据记录的详细信息，包括关联的天气数据
    """
    traffic_data = get_object_or_404(TrafficFlow, pk=pk)

    # 尝试获取同时段的天气数据
    weather_data = None
    try:
        weather_data = WeatherData.objects.filter(
            location=traffic_data.location,
            timestamp__date=traffic_data.timestamp.date(),
        ).first()
    except Exception:
        pass

    # 尝试获取同时段的时段标签
    time_label = None
    try:
        time_label = TimePeriodLabel.objects.filter(
            date=traffic_data.timestamp.date(),
        ).first()
    except Exception:
        pass

    # 获取同点位的预测结果（如有）
    predictions = PredictionResult.objects.filter(
        location=traffic_data.location,
        prediction_time__date=traffic_data.timestamp.date(),
    )[:5]

    context = {
        'data': traffic_data,
        'weather_data': weather_data,
        'time_label': time_label,
        'predictions': predictions,
    }
    return render(request, 'traffic_data/data_detail.html', context)


@login_required
def data_clean_view(request):
    """数据清洗视图
    GET: 显示未清洗数据的统计信息和清洗选项
    POST: 执行数据清洗操作（去除异常值、填补缺失值、标准化处理）
    """
    if request.method == 'POST':
        clean_type = request.POST.get('clean_type', 'all')
        cleaned_count = 0

        try:
            uncleaned_data = TrafficFlow.objects.filter(is_cleaned=False)

            if clean_type == 'outlier' or clean_type == 'all':
                # 异常值处理：将明显异常的流量数据标记并修正
                # 使用IQR方法检测异常值
                from django.db.models import StdDev
                stats = TrafficFlow.objects.aggregate(
                    avg_flow=Avg('total_flow'),
                    std_flow=StdDev('total_flow'),
                )
                avg_flow = stats['avg_flow'] or 0
                std_flow = stats['std_flow'] or 1

                # 将超过3倍标准差的数据截断
                upper_bound = avg_flow + 3 * std_flow
                lower_bound = max(0, avg_flow - 3 * std_flow)

                outliers = uncleaned_data.filter(
                    Q(total_flow__gt=upper_bound) | Q(total_flow__lt=lower_bound)
                )
                for record in outliers:
                    record.total_flow = min(max(record.total_flow, int(lower_bound)), int(upper_bound))
                    record.save()
                    cleaned_count += 1

            if clean_type == 'missing' or clean_type == 'all':
                # 缺失值处理：将null的速度和占有率用均值填充
                avg_values = TrafficFlow.objects.aggregate(
                    avg_speed_val=Avg('avg_speed'),
                    avg_occupancy=Avg('occupancy_rate'),
                )
                if avg_values['avg_speed_val'] is not None:
                    updated = uncleaned_data.filter(avg_speed__isnull=True).update(
                        avg_speed=round(avg_values['avg_speed_val'], 2)
                    )
                    cleaned_count += updated

                if avg_values['avg_occupancy'] is not None:
                    updated = uncleaned_data.filter(occupancy_rate__isnull=True).update(
                        occupancy_rate=round(avg_values['avg_occupancy'], 2)
                    )
                    cleaned_count += updated

            if clean_type == 'negative' or clean_type == 'all':
                # 负值处理：将负数流量值修正为0
                negative_records = uncleaned_data.filter(
                    Q(vehicle_count__lt=0) | Q(pedestrian_count__lt=0) |
                    Q(bicycle_count__lt=0) | Q(motorcycle_count__lt=0) |
                    Q(truck_count__lt=0)
                )
                for record in negative_records:
                    record.vehicle_count = max(0, record.vehicle_count)
                    record.pedestrian_count = max(0, record.pedestrian_count)
                    record.bicycle_count = max(0, record.bicycle_count)
                    record.motorcycle_count = max(0, record.motorcycle_count)
                    record.truck_count = max(0, record.truck_count)
                    record.save()
                    cleaned_count += 1

            # 标记为已清洗
            uncleaned_data.update(is_cleaned=True)

            _log_activity(request.user, 'upload', f'执行数据清洗: 类型={clean_type}，处理{cleaned_count}条', request)
            messages.success(request, f'数据清洗完成！共处理 {cleaned_count} 条记录。')

        except Exception as e:
            messages.error(request, f'数据清洗过程出错: {str(e)}')

        return redirect('data_clean')

    # GET请求：显示清洗统计信息
    total_count = TrafficFlow.objects.count()
    uncleaned_count = TrafficFlow.objects.filter(is_cleaned=False).count()
    cleaned_count = TrafficFlow.objects.filter(is_cleaned=True).count()

    # 数据质量统计
    null_speed_count = TrafficFlow.objects.filter(avg_speed__isnull=True).count()
    null_occupancy_count = TrafficFlow.objects.filter(occupancy_rate__isnull=True).count()
    negative_count = TrafficFlow.objects.filter(
        Q(vehicle_count__lt=0) | Q(pedestrian_count__lt=0) |
        Q(bicycle_count__lt=0) | Q(motorcycle_count__lt=0) |
        Q(truck_count__lt=0)
    ).count()

    context = {
        'total_count': total_count,
        'uncleaned_count': uncleaned_count,
        'cleaned_count': cleaned_count,
        'null_speed_count': null_speed_count,
        'null_occupancy_count': null_occupancy_count,
        'negative_count': negative_count,
    }
    return render(request, 'traffic_data/data_clean.html', context)


@login_required
def data_quality_view(request):
    """数据质量监控仪表板视图
    展示数据完整性、准确性和时效性等质量指标
    """
    now = timezone.now()
    last_24h = now - timedelta(hours=24)
    last_7d = now - timedelta(days=7)

    # 总体数据统计
    total_records = TrafficFlow.objects.count()
    recent_records = TrafficFlow.objects.filter(created_at__gte=last_24h).count()

    # 数据完整性指标
    completeness = {
        'total': total_records,
        'with_speed': TrafficFlow.objects.filter(avg_speed__isnull=False).count(),
        'with_occupancy': TrafficFlow.objects.filter(occupancy_rate__isnull=False).count(),
        'with_confidence': TrafficFlow.objects.filter(confidence__isnull=False).count(),
        'cleaned': TrafficFlow.objects.filter(is_cleaned=True).count(),
    }
    if total_records > 0:
        completeness['speed_rate'] = round(completeness['with_speed'] / total_records * 100, 1)
        completeness['occupancy_rate'] = round(completeness['with_occupancy'] / total_records * 100, 1)
        completeness['confidence_rate'] = round(completeness['with_confidence'] / total_records * 100, 1)
        completeness['clean_rate'] = round(completeness['cleaned'] / total_records * 100, 1)
    else:
        completeness['speed_rate'] = 0
        completeness['occupancy_rate'] = 0
        completeness['confidence_rate'] = 0
        completeness['clean_rate'] = 0

    # 按数据源统计
    source_stats = TrafficFlow.objects.values('source').annotate(
        count=Count('id'),
        avg_flow=Avg('total_flow'),
        avg_confidence=Avg('confidence'),
    ).order_by('-count')

    # 按监测点位统计
    location_stats = TrafficFlow.objects.values('location').annotate(
        count=Count('id'),
        avg_flow=Avg('total_flow'),
        latest=Max('timestamp'),
    ).order_by('-count')[:20]

    # 最近7天每天的数据量趋势
    daily_counts = []
    for i in range(7):
        day = (now - timedelta(days=i)).date()
        count = TrafficFlow.objects.filter(timestamp__date=day).count()
        daily_counts.append({'date': day.strftime('%m-%d'), 'count': count})
    daily_counts.reverse()

    # 天气数据统计
    weather_count = WeatherData.objects.count()
    time_label_count = TimePeriodLabel.objects.count()

    context = {
        'total_records': total_records,
        'recent_records': recent_records,
        'completeness': completeness,
        'source_stats': source_stats,
        'location_stats': location_stats,
        'daily_counts': daily_counts,
        'weather_count': weather_count,
        'time_label_count': time_label_count,
    }
    return render(request, 'traffic_data/data_quality.html', context)
