"""报告生成模块 - 视图函数"""
import os
import json
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, Http404
from django.conf import settings
from django.db.models import Avg, Count, Sum, Max, Min
from django.utils import timezone

from apps.traffic_data.models import TrafficFlow, WeatherData, PredictionResult
from apps.prediction.models import TrainedModel
from apps.users.views import _log_activity


# 报告存储目录
REPORT_DIR = os.path.join(settings.MEDIA_ROOT, 'reports')


@login_required
def report_list_view(request):
    """报告列表视图
    展示已生成的所有报告文件，支持分页和搜索
    """
    # 确保报告目录存在
    os.makedirs(REPORT_DIR, exist_ok=True)

    # 获取报告文件列表
    reports = []
    if os.path.exists(REPORT_DIR):
        for filename in sorted(os.listdir(REPORT_DIR), reverse=True):
            filepath = os.path.join(REPORT_DIR, filename)
            if os.path.isfile(filepath):
                file_stat = os.stat(filepath)
                reports.append({
                    'filename': filename,
                    'size': round(file_stat.st_size / 1024, 1),  # KB
                    'created_time': datetime.fromtimestamp(file_stat.st_mtime),
                    'file_type': filename.rsplit('.', 1)[-1].upper() if '.' in filename else '未知',
                })

    # 按文件类型过滤
    file_type = request.GET.get('file_type', '')
    if file_type:
        reports = [r for r in reports if r['file_type'].lower() == file_type.lower()]

    # 搜索
    search = request.GET.get('search', '')
    if search:
        reports = [r for r in reports if search.lower() in r['filename'].lower()]

    # 手动分页
    page_size = 20
    page_number = int(request.GET.get('page', 1))
    total_pages = max(1, (len(reports) + page_size - 1) // page_size)
    page_number = min(max(1, page_number), total_pages)
    start_idx = (page_number - 1) * page_size
    end_idx = start_idx + page_size
    page_reports = reports[start_idx:end_idx]

    context = {
        'reports': page_reports,
        'total_count': len(reports),
        'page_number': page_number,
        'total_pages': total_pages,
        'has_previous': page_number > 1,
        'has_next': page_number < total_pages,
        'previous_page': page_number - 1,
        'next_page': page_number + 1,
        'file_type': file_type,
        'search': search,
    }
    return render(request, 'reports/report_list.html', context)


@login_required
def generate_report_view(request):
    """报告生成视图
    GET: 显示报告生成配置表单
    POST: 根据参数生成预测分析报告（PDF/Excel/CSV）
    """
    if request.method == 'POST':
        report_type = request.POST.get('report_type', 'csv')  # csv, excel, pdf
        date_from = request.POST.get('date_from', '')
        date_to = request.POST.get('date_to', '')
        location = request.POST.get('location', '')
        report_name = request.POST.get('report_name', '').strip()

        # 确定日期范围
        now = timezone.now()
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d') if date_from else (now - timedelta(days=7))
            end_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1) if date_to else now
        except ValueError:
            start_date = now - timedelta(days=7)
            end_date = now

        # 查询数据
        traffic_data = TrafficFlow.objects.filter(
            timestamp__gte=start_date,
            timestamp__lte=end_date,
        )
        if location:
            traffic_data = traffic_data.filter(location__icontains=location)

        prediction_data = PredictionResult.objects.filter(
            prediction_time__gte=start_date,
            prediction_time__lte=end_date,
        )
        if location:
            prediction_data = prediction_data.filter(location__icontains=location)

        # 数据统计
        traffic_stats = traffic_data.aggregate(
            total_records=Count('id'),
            avg_flow=Avg('total_flow'),
            max_flow=Max('total_flow'),
            min_flow=Min('total_flow'),
            sum_total_flow=Sum('total_flow'),
            avg_speed=Avg('avg_speed'),
        )
        prediction_stats = prediction_data.aggregate(
            total_predictions=Count('id'),
            avg_mae=Avg('mae'),
            avg_rmse=Avg('rmse'),
            avg_mape=Avg('mape'),
        )

        if not report_name:
            report_name = f'traffic_report_{now.strftime("%Y%m%d_%H%M%S")}'

        # 确保报告目录存在
        os.makedirs(REPORT_DIR, exist_ok=True)

        try:
            if report_type == 'csv':
                # 生成CSV报告
                filename = f'{report_name}.csv'
                filepath = os.path.join(REPORT_DIR, filename)

                import csv
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    # 写入报告头信息
                    writer.writerow(['交通流量预测分析报告'])
                    writer.writerow([f'生成时间: {now.strftime("%Y-%m-%d %H:%M:%S")}'])
                    writer.writerow([f'数据范围: {start_date.strftime("%Y-%m-%d")} 至 {end_date.strftime("%Y-%m-%d")}'])
                    writer.writerow([])

                    # 写入统计摘要
                    writer.writerow(['=== 数据统计摘要 ==='])
                    writer.writerow(['总记录数', traffic_stats['total_records'] or 0])
                    writer.writerow(['平均流量', round(traffic_stats['avg_flow'] or 0, 1)])
                    writer.writerow(['最大流量', traffic_stats['max_flow'] or 0])
                    writer.writerow(['最小流量', traffic_stats['min_flow'] or 0])
                    writer.writerow(['平均速度(km/h)', round(traffic_stats['avg_speed'] or 0, 1)])
                    writer.writerow([])

                    # 写入预测统计
                    writer.writerow(['=== 预测结果统计 ==='])
                    writer.writerow(['预测总数', prediction_stats['total_predictions'] or 0])
                    writer.writerow(['平均MAE', round(prediction_stats['avg_mae'] or 0, 2)])
                    writer.writerow(['平均RMSE', round(prediction_stats['avg_rmse'] or 0, 2)])
                    writer.writerow(['平均MAPE(%)', round(prediction_stats['avg_mape'] or 0, 2)])
                    writer.writerow([])

                    # 写入详细数据
                    writer.writerow(['=== 交通流量明细数据 ==='])
                    writer.writerow([
                        '时间戳', '监测点位', '总流量', '车辆数', '行人数',
                        '自行车数', '摩托车数', '卡车数', '平均速度', '数据源'
                    ])
                    for record in traffic_data[:1000]:  # 限制最多1000条
                        writer.writerow([
                            record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                            record.location,
                            record.total_flow,
                            record.vehicle_count,
                            record.pedestrian_count,
                            record.bicycle_count,
                            record.motorcycle_count,
                            record.truck_count,
                            record.avg_speed or '',
                            record.get_source_display(),
                        ])

                messages.success(request, f'CSV报告已生成: {filename}')

            elif report_type == 'excel':
                # 生成Excel报告
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                except ImportError:
                    messages.error(request, '服务器缺少openpyxl库，无法生成Excel报告。')
                    return redirect('generate_report')

                filename = f'{report_name}.xlsx'
                filepath = os.path.join(REPORT_DIR, filename)

                wb = openpyxl.Workbook()

                # 摘要工作表
                ws_summary = wb.active
                ws_summary.title = '报告摘要'
                header_font = Font(bold=True, size=14)
                ws_summary.append(['交通流量预测分析报告'])
                ws_summary['A1'].font = header_font
                ws_summary.append([f'生成时间: {now.strftime("%Y-%m-%d %H:%M:%S")}'])
                ws_summary.append([f'数据范围: {start_date.strftime("%Y-%m-%d")} 至 {end_date.strftime("%Y-%m-%d")}'])
                ws_summary.append([])
                ws_summary.append(['指标', '数值'])
                ws_summary.append(['总记录数', traffic_stats['total_records'] or 0])
                ws_summary.append(['平均流量', round(traffic_stats['avg_flow'] or 0, 1)])
                ws_summary.append(['最大流量', traffic_stats['max_flow'] or 0])
                ws_summary.append(['最小流量', traffic_stats['min_flow'] or 0])
                ws_summary.append(['平均速度(km/h)', round(traffic_stats['avg_speed'] or 0, 1)])
                ws_summary.append(['预测总数', prediction_stats['total_predictions'] or 0])
                ws_summary.append(['平均MAE', round(prediction_stats['avg_mae'] or 0, 2)])
                ws_summary.append(['平均RMSE', round(prediction_stats['avg_rmse'] or 0, 2)])
                ws_summary.append(['平均MAPE(%)', round(prediction_stats['avg_mape'] or 0, 2)])

                # 详细数据工作表
                ws_detail = wb.create_sheet('流量明细')
                headers = ['时间戳', '监测点位', '总流量', '车辆数', '行人数',
                           '自行车数', '摩托车数', '卡车数', '平均速度', '数据源']
                ws_detail.append(headers)
                for cell in ws_detail[1]:
                    cell.font = Font(bold=True)

                for record in traffic_data[:5000]:
                    ws_detail.append([
                        record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                        record.location,
                        record.total_flow,
                        record.vehicle_count,
                        record.pedestrian_count,
                        record.bicycle_count,
                        record.motorcycle_count,
                        record.truck_count,
                        record.avg_speed or 0,
                        record.get_source_display(),
                    ])

                # 预测结果工作表
                ws_pred = wb.create_sheet('预测结果')
                pred_headers = ['模型类型', '版本', '点位', '预测时间', '预测流量',
                                '实际流量', 'MAE', 'RMSE', 'MAPE(%)']
                ws_pred.append(pred_headers)
                for cell in ws_pred[1]:
                    cell.font = Font(bold=True)

                for pred in prediction_data[:5000]:
                    ws_pred.append([
                        pred.get_model_type_display(),
                        pred.model_version,
                        pred.location,
                        pred.prediction_time.strftime('%Y-%m-%d %H:%M:%S'),
                        pred.predicted_flow,
                        pred.actual_flow or '',
                        pred.mae or '',
                        pred.rmse or '',
                        pred.mape or '',
                    ])

                wb.save(filepath)
                messages.success(request, f'Excel报告已生成: {filename}')

            elif report_type == 'pdf':
                # 生成PDF报告（简单文本格式，使用reportlab时可替换）
                filename = f'{report_name}.pdf'
                filepath = os.path.join(REPORT_DIR, filename)

                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib.units import cm
                    from reportlab.pdfgen import canvas
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont

                    c = canvas.Canvas(filepath, pagesize=A4)
                    width, height = A4

                    # 尝试注册中文字体
                    try:
                        pdfmetrics.registerFont(TTFont('SimHei', 'SimHei.ttf'))
                        font_name = 'SimHei'
                    except Exception:
                        font_name = 'Helvetica'

                    # 标题
                    c.setFont(font_name, 18)
                    c.drawString(2 * cm, height - 3 * cm, 'Traffic Flow Prediction Report')

                    c.setFont(font_name, 10)
                    y = height - 4.5 * cm
                    c.drawString(2 * cm, y, f'Generated: {now.strftime("%Y-%m-%d %H:%M:%S")}')
                    y -= 0.6 * cm
                    c.drawString(2 * cm, y, f'Date Range: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}')

                    y -= 1.5 * cm
                    c.setFont(font_name, 12)
                    c.drawString(2 * cm, y, 'Data Summary')

                    c.setFont(font_name, 10)
                    y -= 0.8 * cm
                    summary_lines = [
                        f'Total Records: {traffic_stats["total_records"] or 0}',
                        f'Average Flow: {round(traffic_stats["avg_flow"] or 0, 1)}',
                        f'Max Flow: {traffic_stats["max_flow"] or 0}',
                        f'Min Flow: {traffic_stats["min_flow"] or 0}',
                        f'Average Speed: {round(traffic_stats["avg_speed"] or 0, 1)} km/h',
                        '',
                        f'Total Predictions: {prediction_stats["total_predictions"] or 0}',
                        f'Average MAE: {round(prediction_stats["avg_mae"] or 0, 2)}',
                        f'Average RMSE: {round(prediction_stats["avg_rmse"] or 0, 2)}',
                        f'Average MAPE: {round(prediction_stats["avg_mape"] or 0, 2)}%',
                    ]
                    for line in summary_lines:
                        c.drawString(2.5 * cm, y, line)
                        y -= 0.5 * cm

                    c.save()
                    messages.success(request, f'PDF报告已生成: {filename}')

                except ImportError:
                    # reportlab未安装，生成纯文本替代
                    filename = f'{report_name}.txt'
                    filepath = os.path.join(REPORT_DIR, filename)
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write('交通流量预测分析报告\n')
                        f.write('=' * 50 + '\n')
                        f.write(f'生成时间: {now.strftime("%Y-%m-%d %H:%M:%S")}\n')
                        f.write(f'数据范围: {start_date.strftime("%Y-%m-%d")} 至 {end_date.strftime("%Y-%m-%d")}\n\n')
                        f.write('数据统计摘要\n')
                        f.write('-' * 30 + '\n')
                        f.write(f'总记录数: {traffic_stats["total_records"] or 0}\n')
                        f.write(f'平均流量: {round(traffic_stats["avg_flow"] or 0, 1)}\n')
                        f.write(f'最大流量: {traffic_stats["max_flow"] or 0}\n')
                        f.write(f'最小流量: {traffic_stats["min_flow"] or 0}\n')
                        f.write(f'平均速度: {round(traffic_stats["avg_speed"] or 0, 1)} km/h\n\n')
                        f.write('预测结果统计\n')
                        f.write('-' * 30 + '\n')
                        f.write(f'预测总数: {prediction_stats["total_predictions"] or 0}\n')
                        f.write(f'平均MAE: {round(prediction_stats["avg_mae"] or 0, 2)}\n')
                        f.write(f'平均RMSE: {round(prediction_stats["avg_rmse"] or 0, 2)}\n')
                        f.write(f'平均MAPE: {round(prediction_stats["avg_mape"] or 0, 2)}%\n')
                    messages.warning(request, f'reportlab库未安装，已生成文本格式报告: {filename}')

            else:
                messages.error(request, '不支持的报告格式。')
                return redirect('generate_report')

            # 记录导出活动
            _log_activity(request.user, 'export', f'生成报告: {filename}', request)

        except Exception as e:
            messages.error(request, f'报告生成失败: {str(e)}')
            return redirect('generate_report')

        return redirect('report_list')

    # GET请求：显示报告生成表单
    locations = TrafficFlow.objects.values_list('location', flat=True).distinct()[:50]
    context = {
        'locations': locations,
        'data_count': TrafficFlow.objects.count(),
        'prediction_count': PredictionResult.objects.count(),
    }
    return render(request, 'reports/generate_report.html', context)


@login_required
def report_download_view(request, filename):
    """报告下载视图
    根据文件名下载已生成的报告文件
    """
    filepath = os.path.join(REPORT_DIR, filename)

    # 安全检查：防止目录遍历攻击
    if not os.path.abspath(filepath).startswith(os.path.abspath(REPORT_DIR)):
        raise Http404('文件不存在')

    if not os.path.exists(filepath):
        raise Http404('文件不存在')

    # 确定Content-Type
    content_types = {
        '.csv': 'text/csv',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.pdf': 'application/pdf',
        '.txt': 'text/plain',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    }
    ext = os.path.splitext(filename)[1].lower()
    content_type = content_types.get(ext, 'application/octet-stream')

    with open(filepath, 'rb') as f:
        response = HttpResponse(f.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
